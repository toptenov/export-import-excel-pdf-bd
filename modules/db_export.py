import datetime
import os

import openpyxl
from fpdf import FPDF


def create_result_directories():
    if not os.path.exists("results"):
        os.mkdir("results")
        print('/result directory has been successfully created')
    if not os.path.exists("results/Excel"):
        os.mkdir("results/Excel")
        print('/results/Excel directory has been successfully created')
    if not os.path.exists("results/PDF"):
        os.mkdir("results/PDF")
        print('/results/PDF directory has been successfully created')


def get_all_users_from_db(sql):
    users_selection = sql.execute(
        '''
        SELECT u.id, first_name, second_name, patronymic, region_name, city_name, phone, email
        FROM users u
        JOIN regions r ON r.id = u.region_id
        JOIN cities c ON c.id = u.city_id
        '''
    )
    return users_selection


def export_data_to_excel(sql):
    book = openpyxl.Workbook()
    users_sheet = book.active
    users_sheet.title = 'users'

    users_selection = get_all_users_from_db(sql)

    # Fill excel header:
    excel_header = ('id', 'first_name', 'second_name', 'patronymic', 'region_name', 'city_name', 'phone', 'email')
    for i in range(1, len(excel_header) + 1):
        users_sheet.cell(row=1, column=i).value = excel_header[i - 1]

    # Fill excel with data from DB:
    for i, row in enumerate(users_selection):
        i += 1
        for j, value in enumerate(row):
            j += 1
            users_sheet.cell(row=i + 1, column=j).value = value

    result_path = 'results/Excel/' + datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + ' Result.xlsx'
    book.save(filename=result_path)
    print(f'Users from database have been successfully exported to Excel: {result_path}')


def export_data_to_pdf(sql):
    pdf = FPDF()
    font_file_path = os.getcwd() + '/fonts/DejaVuSans'
    pdf.add_font(family='DejaVuSans', style='', fname=font_file_path+'.ttf', uni=True)
    pdf.add_font(family='DejaVuSans', style='B', fname=font_file_path+'-Bold.ttf', uni=True)

    users_selection = get_all_users_from_db(sql)

    # Fill PDF with data from DB:
    for i, row in enumerate(users_selection):
        pdf.add_page()
        pdf.set_font(family='DejaVuSans', style='B', size=25)
        pdf.cell(w=190, h=17, txt=row[2] + ' ' + row[1] + ' ' + row[3], ln=1, align="C")
        pdf.set_font(family='DejaVuSans', style='', size=10)
        pdf.cell(w=190, h=5, txt='Телефон: ' + row[6], ln=1, align="L")
        pdf.cell(w=190, h=5, txt='Почта: ' + row[7], ln=1, align="L")
        pdf.cell(w=190, h=5, txt='Регион: ' + row[4], ln=1, align="L")
        pdf.cell(w=190, h=5, txt='Город: ' + row[5], ln=1, align="L")

    result_path = 'results/PDF/' + datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S') + ' Result.pdf'
    pdf.output(result_path)
    print(f'Users from database have been successfully exported to PDF: {result_path}')
